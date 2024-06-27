from openpyxl import Workbook

# Datos de ejemplo para equipos de Pokémon
teams_data = [
    {"id": 1, "name": "Team 1", "pokemon1": "Pikachu", "pokemon2": "Charizard", "pokemon3": "Blastoise",
     "pokemon4": "Venusaur", "pokemon5": "Gengar", "pokemon6": "Dragonite"},
    {"id": 2, "name": "Team 2", "pokemon1": "Jolteon", "pokemon2": "Alakazam", "pokemon3": "Snorlax",
     "pokemon4": "Gyarados", "pokemon5": "Exeggutor", "pokemon6": "Mewtwo"},
    {"id": 3, "name": "Team 3", "pokemon1": "Blaziken", "pokemon2": "Tyranitar", "pokemon3": "Metagross",
     "pokemon4": "Salamence", "pokemon5": "Garchomp", "pokemon6": "Volcarona"}
]

# Datos de ejemplo para torneos
tournaments_data = [
    {"id": 1, "name": "Tournament A", "phase": "group"},
    {"id": 2, "name": "Tournament B", "phase": "double elimination"},
    {"id": 3, "name": "Tournament C", "phase": "group"}
]

# Datos de ejemplo para partidas de torneo
matches_data = [
    {"id": 1, "tournament_id": 1, "player1": "Team 1", "player2": "Team 2", "result": "Team 1 wins"},
    {"id": 2, "tournament_id": 1, "player1": "Team 3", "player2": "Team 1", "result": "Team 3 wins"},
    {"id": 3, "tournament_id": 2, "player1": "Team 2", "player2": "Team 3", "result": "Team 2 wins"}
]

# Datos de ejemplo para usuarios
users_data = [
    {"id": 1, "username": "user1", "password": "password1"},
    {"id": 2, "username": "user2", "password": "password2"},
    {"id": 3, "username": "user3", "password": "password3"}
]

# Datos de ejemplo para Pokémon
pokemon_data = [
    {"id": 1, "name": "Pikachu", "type1": "Electric", "type2": "", "hp": 35, "attack": 55, "defense": 40,
     "special_attack": 50, "special_defense": 50, "speed": 90, "ability": "Static", "move1": "Thunderbolt",
     "move2": "Quick Attack", "move3": "Iron Tail", "move4": "Volt Tackle", "item": "Light Ball"},
    {"id": 2, "name": "Charizard", "type1": "Fire", "type2": "Flying", "hp": 78, "attack": 84, "defense": 78,
     "special_attack": 109, "special_defense": 85, "speed": 100, "ability": "Blaze", "move1": "Flamethrower",
     "move2": "Air Slash", "move3": "Dragon Claw", "move4": "Solar Beam", "item": "Charizardite Y"},
    {"id": 3, "name": "Blastoise", "type1": "Water", "type2": "", "hp": 79, "attack": 83, "defense": 100,
     "special_attack": 85, "special_defense": 105, "speed": 78, "ability": "Torrent", "move1": "Hydro Pump",
     "move2": "Ice Beam", "move3": "Aqua Tail", "move4": "Rapid Spin", "item": "Blastoisinite"},
    {"id": 4, "name": "Venusaur", "type1": "Grass", "type2": "Poison", "hp": 80, "attack": 82, "defense": 83,
     "special_attack": 100, "special_defense": 100, "speed": 80, "ability": "Overgrow", "move1": "Solar Beam",
     "move2": "Sludge Bomb", "move3": "Earthquake", "move4": "Synthesis", "item": "Venusaurite"},
    {"id": 5, "name": "Gengar", "type1": "Ghost", "type2": "Poison", "hp": 60, "attack": 65, "defense": 60,
     "special_attack": 130, "special_defense": 75, "speed": 110, "ability": "Levitate", "move1": "Shadow Ball",
     "move2": "Sludge Bomb", "move3": "Focus Blast", "move4": "Thunderbolt", "item": "Gengarite"},
    {"id": 6, "name": "Dragonite", "type1": "Dragon", "type2": "Flying", "hp": 91, "attack": 134, "defense": 95,
     "special_attack": 100, "special_defense": 100, "speed": 80, "ability": "Inner Focus", "move1": "Outrage",
     "move2": "Dragon Dance", "move3": "Earthquake", "move4": "Fire Punch", "item": "Leftovers"}
]

# Crear un nuevo libro de trabajo (Workbook)
wb = Workbook()

# Hoja para equipos
ws_teams = wb.active
ws_teams.title = "Teams"
ws_teams.append(["id", "name", "pokemon1", "pokemon2", "pokemon3", "pokemon4", "pokemon5", "pokemon6"])
for team in teams_data:
    ws_teams.append([team["id"], team["name"], team["pokemon1"], team["pokemon2"], team["pokemon3"],
                     team["pokemon4"], team["pokemon5"], team["pokemon6"]])

# Hoja para torneos
ws_tournaments = wb.create_sheet("Tournaments")
ws_tournaments.append(["id", "name", "phase"])
for tournament in tournaments_data:
    ws_tournaments.append([tournament["id"], tournament["name"], tournament["phase"]])

# Hoja para partidas de torneo
ws_matches = wb.create_sheet("Matches")
ws_matches.append(["id", "tournament_id", "player1", "player2", "result"])
for match in matches_data:
    ws_matches.append([match["id"], match["tournament_id"], match["player1"], match["player2"], match["result"]])

# Hoja para usuarios
ws_users = wb.create_sheet("Users")
ws_users.append(["id", "username", "password"])
for user in users_data:
    ws_users.append([user["id"], user["username"], user["password"]])

# Hoja para Pokémon
ws_pokemon = wb.create_sheet("Pokemon")
ws_pokemon.append(["id", "name", "type1", "type2", "hp", "attack", "defense", "special_attack", "special_defense",
                   "speed", "ability", "move1", "move2", "move3", "move4", "item"])
for poke in pokemon_data:
    ws_pokemon.append([poke["id"], poke["name"], poke["type1"], poke["type2"], poke["hp"], poke["attack"],
                       poke["defense"], poke["special_attack"], poke["special_defense"], poke["speed"],
                       poke["ability"], poke["move1"], poke["move2"], poke["move3"], poke["move4"], poke["item"]])

# Guardar el archivo
wb.save("pokemon_app_data.xlsx")
print("Archivo pokemon_app_data.xlsx creado correctamente.")
